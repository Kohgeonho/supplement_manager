import pandas as pd
import tkinter.ttk as ttk
import tkinter
import tkinter.messagebox
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
import re
from pandastable import Table

# Tool 창 설정
tool = Tk()
tool.title("보충보급 관리 프로그램")
tool.geometry("450x350+600+300")    # 가로 * 세로 + x좌표 + y좌표
tool.resizable(False, False)    # 높이, 너비 값 변경 불가
tool.iconphoto(True, PhotoImage(file='icon.gif'))


# 정보 입력
dataframe = LabelFrame(tool, text="정보 입력", font="맑은고딕 18 bold")
dataframe.pack(side="top", pady=10, ipady=5)


# 결과 조회
resultframe = LabelFrame(tool, text="결과 조회", font="맑은고딕 18 bold")
resultframe.pack(side="bottom", pady=30, ipady=5)


# 입대 년 월
dateframe = LabelFrame(dataframe, text="입대 년 월", font="맑은고딕 12")
dateframe.grid(row=0, column=0, padx=10, pady=5)

years = [str(i) + "년" for i in range(2020, 2030)]
yearbox = ttk.Combobox(dateframe, width=8, height=10, values=years, state="readonly", font="맑은고딕 12")
yearbox.current(0)    # 0번째 인덱스 값 선택
yearbox.pack(side="left", padx=3, pady=5)

months = [str(i) + "월" for i in range(1, 13)]
monthbox = ttk.Combobox(dateframe, width=8, height=12, values=months, state="readonly", font="맑은고딕 12")
monthbox.current(0)
monthbox.pack(side="right", padx=3, pady=5)

def msgbox(title, info):
    tkinter.messagebox.showinfo(title, info)
def errorbox(title, info):
    tkinter.messagebox.showerror(title, info)
def questionbox(title, info):
    return tkinter.messagebox.askokcancel(title, info)


# 최초 인사 명령 결과 추가
def HR_result():

    # 파일 선택
    def select_file():
        files = filedialog.askopenfilename(title="파일을 선택하세요", \
                                           filetypes=(("Excel 파일", "*.xls*"), \
                                                      ("Cell 파일", "*.cell"), \
                                                      ("모든 파일", "*.*")), \
                                           initialdir="./")    # 현재 경로를 보여줌
        hr = pd.read_excel(files)

        unit_nums = [re.compile('^(.*) 부 \(이상(.*)명\)$') \
                       .findall(string.strip())
                     for string in hr['Unnamed: 6'].dropna(how='any')]

        units = []
        for unit_num in unit_nums:
            unit, num = unit_num[0]
            units += [unit] * int(num)

        df = pd.DataFrame()
        df['군번'] = hr['Unnamed: 26']
        df['이름'] = hr['Unnamed: 33']
        hr_sliced = df.dropna(how='any')
        
        unattached = len(hr_sliced) - len(units)
        hr_sliced['부대'] = units + ['소속없음'] * unattached
        return hr_sliced[['부대', '군번', '이름']]

    # 인사 정보 엑셀 시트에 저장
    def save_data(df):
        year = yearbox.get()
        month = monthbox.get()

        wb = load_workbook('member_info.xlsx')
        overwrite = True
        if year+month in wb.sheetnames:
            overwrite = questionbox("중복된 시트", f"{year+month}에 해당하는 정보가 이미 존재합니다. 바꾸시겠습니까?")

        if overwrite == True:
            writer = pd.ExcelWriter('member_info.xlsx', mode='a', if_sheet_exists='replace')
            df.to_excel(writer, sheet_name=year+month, index=False)
            writer.save()
            msgbox("최초 인사 명령 결과", "완료되었습니다.")

    try:
        save_data(select_file())
    except PermissionError:
        errorbox("Permission Error", "접근 권한이 없습니다. \n 파일을 닫고 다시 진행해주십시오.")
    except FileNotFoundError:
        errorbox("File Not Found Error", "파일 선택이 취소되었습니다.")


# 최초 인사 명령 결과 버튼
btn1 = Button(dataframe, width=20, height=2, text="최초 인사 명령 결과", font="맑은고딕 12", command=HR_result)
btn1.grid(row=0, column=1, padx=10, pady=5)


# 피복 사이즈 정보 추가
def sizeInfo():
    wb = load_workbook('member_info.xlsx')
    sheets = wb.sheetnames

    try:
        files = filedialog.askopenfilename(title="파일을 선택하세요", \
                                        filetypes=(("Excel 파일", "*.xls*"), \
                                                    ("Cell 파일", "*.cell"), \
                                                    ("모든 파일", "*.*")), \
                                        initialdir="./")    # 현재 경로를 보여줌
        size_info = pd.read_excel(files, header=3)

        for sheet in sheets:
            member = pd.read_excel('member_info.xlsx', sheet_name=sheet)

            joined = member[['군번', '이름', '부대']].set_index('군번') \
                           .join(size_info[['군번', '런닝', '팬티', '슬리퍼']].set_index('군번'), on='군번')
            writer = pd.ExcelWriter('member_info.xlsx', mode='a', if_sheet_exists='replace')
            joined.to_excel(writer, sheet_name=sheet, index=True)
            writer.save()
            
        msgbox("피복 사이즈 정보", "완료되었습니다.")

    except PermissionError:
        errorbox("Permission Error", "접근 권한이 없습니다. \n 파일을 닫고 다시 진행해주십시오.")
    except FileNotFoundError:
        errorbox("File Not Found Error", "파일 선택이 취소되었습니다.")


# 피복 사이즈 정보 버튼
btn2 = Button(dataframe, width=20, height=2, text="피복 사이즈 정보", font="맑은고딕 12", command=sizeInfo)
btn2.grid(row=1, column=0, padx=10, pady=5)


# 최종 부대 분류 결과 버튼
btn3 = Button(dataframe, width=20, height=2, text="최종 부대 분류 결과", font="맑은고딕 12")
btn3.grid(row=1, column=1, padx=10, pady=5)


# 품목별 인원 조회
def show_itemdata():
    item_data = Toplevel(tool)
    item_data.title("품목별 조회")
    item_data.geometry("600x400+200+100")
        
    f = Frame(item_data)
    f.pack(fill=BOTH,expand=1)

    df = pd.read_excel('member_info.xlsx', sheet_name=None)
    items = {'런닝': lambda x:int(x[1:]), 
            '팬티': lambda x:int(x[1:]), 
            '슬리퍼': lambda x:int(x)}
    unit_total = {}

    for item in items:
        unit_total[item] = pd.DataFrame(columns=['품목', '사이즈'])

        for month in df.keys():
            unit_df = df[month][item].value_counts(sort=False) \
                                     .convert_dtypes() \
                                     .rename_axis('사이즈') \
                                     .reset_index(name=month)
            sort_value = unit_df['사이즈'].apply(items[item])
            label = [item] * len(unit_df)

            unit_df['품목'] = label
            unit_df['sort'] = sort_value
            unit_df = unit_df.sort_values(by='sort')

            unit_total[item] = pd.merge(unit_total[item], 
                                        unit_df[['품목', '사이즈', month]], 
                                        how='outer', 
                                        on=['품목', '사이즈'])

    pt = Table(f, dataframe=pd.concat(unit_total))
    pt.show()
    item_data.mainloop()


# 부대별 인원 조회
class MyTable(Table):
    """Custom table class inherits from Table. You can then override required methods"""
    def __init__(self, parent=None, **kwargs):
        Table.__init__(self, parent, **kwargs)
        return

    def show_specific(self, col, value):
        w = Toplevel(self.parentframe)
        w.geometry("600x400+200+100")
        
        f = Frame(w)
        f.pack(fill=BOTH,expand=1)

        xl = pd.ExcelFile('member_info.xlsx')
        sheet = xl.sheet_names[col-1]
        df = pd.read_excel('member_info.xlsx', sheet_name=sheet)
        df = df[df['부대'] == value]
        w.title(value + " / " + sheet)

        pt = Table(f, dataframe=df)
        pt.show()
        return

    def popupMenu(self, event, rows=None, cols=None, outside=None):
        popupmenu = Menu(self, tearoff = 0)
        def popupFocusOut(event):
            popupmenu.unpost()

        row = self.get_row_clicked(event)
        col = self.get_col_clicked(event)
        value = self.model.getValueAt(row, 0)

        popupmenu.add_command(label="자세히", command= lambda: self.show_specific(col, value))

        popupmenu.bind("<FocusOut>", popupFocusOut)
        popupmenu.focus_set()
        popupmenu.post(event.x_root, event.y_root)
        return popupmenu

def show_unitdata():

    unit_data = Toplevel(tool)
    unit_data.title("부대 전체 조회")
    unit_data.geometry("600x400+200+100")
        
    f = Frame(unit_data)
    f.pack(fill=BOTH,expand=1)

    df = pd.read_excel('member_info.xlsx', sheet_name=None)
    unit_total = pd.DataFrame(columns=['부대'])
    for month in df.keys():
        unit_df = df[month]['부대'].value_counts(sort=False).convert_dtypes().rename_axis('부대').reset_index(name=month)
        unit_total = pd.merge(unit_total,unit_df, how='outer', on='부대')
    pt = MyTable(f, dataframe=unit_total)
    pt.show()
    unit_data.mainloop()


# 품목별 조회 버튼
btn4 = Button(resultframe, width=20, height=2, text="품목별 조회", font="맑은고딕 12", command=show_itemdata)
btn4.grid(row=0, column=0, padx=10, pady=5)


# 부대별 조회 버튼
btn5 = Button(resultframe, width=20, height=2, text="부대별 조회", font="맑은고딕 12", command=show_unitdata)
btn5.grid(row=0, column=1, padx=10, pady=5)

tool.mainloop()